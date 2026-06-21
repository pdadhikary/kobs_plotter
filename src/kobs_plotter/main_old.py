import sys
from pathlib import Path
import numpy as np


def pick_file_dialog() -> Path | None:
    """Open a native file dialog and return the selected path, or None if cancelled."""
    try:
        import tkinter as tk
        from tkinter import filedialog

        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        path = filedialog.askopenfilename(
            title="Select Excel file",
            filetypes=[("Excel files", "*.xlsx *.xls *.xlsm"), ("All files", "*.*")],
        )
        root.destroy()
        return Path(path) if path else None
    except Exception as e:
        print(f"[error] Could not open file dialog: {e}", file=sys.stderr)
        return None


def get_sheet_names(data_path: Path) -> list[str]:
    try:
        from openpyxl import load_workbook

        wb = load_workbook(data_path)
        sheet_names = wb.sheetnames
        wb.close()
        return sheet_names
    except Exception as e:
        print(f"[error] Could not retrieve sheetnames: {e}", file=sys.stderr)

    return []


def exponential_model(x, a, b, k):
    return b - a * np.exp(-k * x)


def analyze(data_path: Path, sheet_name: str, x_variable: str, y_variable: str) -> None:
    import matplotlib.pyplot as plt
    import pandas as pd
    from scipy.optimize import curve_fit

    df = pd.read_excel(data_path, sheet_name=sheet_name)

    if x_variable not in df.columns or y_variable not in df.columns:
        sys.exit(
            f"[error] X ({x_variable}) and Y ({y_variable}) variable not found in file."
        )

    x = np.array(df[x_variable])
    y = np.array(df[y_variable])

    popt, pcov = curve_fit(exponential_model, x, y, method="lm")

    a_opt, b_opt, k_opt = popt
    perr = np.sqrt(np.diag(pcov))

    n = len(x)
    p = 3
    residuals = y - exponential_model(x, *popt)
    ss_res = np.sum(residuals**2)
    ss_tot = np.sum((y - np.mean(y)) ** 2)
    r2 = 1 - ss_res / ss_tot
    r2_adj = 1 - (1 - r2) * ((n - 1) / (n - p - 1))

    result_string = (
        "Result:\n"
        f"Equation: B - A * exp(-k * x)\n"
        f"A = {a_opt:0.4f} +/- {perr[0]:.04f}\n"
        f"B = {b_opt:0.4f} +/- {perr[1]:.04f}\n"
        f"k = {k_opt:0.4f} +/- {perr[2]:.04f}\n"
        f"R-Square (COD) = {r2:0.5f}\n"
        f"Adj. R-Square = {r2_adj: 0.5f}"
    )

    print("\n" + result_string)

    print("\nPlot Settings")
    plot_title = input("Plot Title: ").strip()
    x_label = input("Plot X-axis Label: ").strip()
    y_label = input("Plot Y-axis Label: ").strip()
    x_axis_scale = eval(input("X-axis Scale factor (default=1.0): ").strip() or "1.0")

    plt.scatter(x * x_axis_scale, y, c="black")
    x_plot = np.linspace(min(x), max(x), 1_000)
    plt.plot(x_plot * x_axis_scale, exponential_model(x_plot, *popt), c="red")

    plt.text(
        0.60,
        0.05,
        result_string,
        transform=plt.gca().transAxes,
        fontsize=9,
        verticalalignment="bottom",
        horizontalalignment="left",
        bbox=dict(boxstyle="round", facecolor="white", alpha=0.8),
    )

    plt.title(plot_title)
    plt.xlabel(x_label)
    plt.ylabel(y_label)

    plt.tight_layout()
    plt.show()


def main():
    data_path = input(
        "Provide the path to the Excel file and press <ENTER> (Leave empty to open file dialog): "
    ).strip()

    if data_path:
        p = Path(data_path)
        if not p.exists() or not p.is_file():
            sys.exit(f"[error] File not found: {p}")
        data_path = p
    else:
        data_path = None
        while data_path is None:
            data_path = pick_file_dialog()
            if data_path is None:
                input("No file selected. Press <ENTER> to try again...")

    print(f'Analyzing file: "{data_path}"...\n')

    sheet_names = get_sheet_names(data_path)

    print("Sheets:")
    for i, sheet in enumerate(sheet_names, start=1):
        print(f"\t{i} - {sheet}")
    sheet_idx = input("Select sheet you want to analyze (default=1): ").strip()
    sheet_idx = int(sheet_idx) if sheet_idx.isnumeric() else 1
    print()

    if sheet_idx > len(sheet_names):
        sys.exit("[error] Sheet index out of range.")

    sheet_name = sheet_names[sheet_idx - 1]

    print(f"Reading sheet: {sheet_name}\n")

    x_variable = input("X column name (default='t'): ").strip()
    x_variable = x_variable if x_variable else "t"

    print()

    y_varibale = input("Y column name (default='C'): ").strip()
    y_varibale = y_varibale if y_varibale else "C"

    analyze(data_path, sheet_name, x_variable, y_varibale)
